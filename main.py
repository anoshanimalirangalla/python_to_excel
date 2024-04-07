##### Assignment 9 ################

# import the necessary packages
import urllib.request
import requests
import json
import openpyxl
import openpyxl.utils 
from openpyxl.chart import Reference, LineChart
import pandas as pd

# get the data from API
API_key = 'b70d1042b2355f2d20df93fe189c57c3'
cities = ['Sudbury', 'Toronto', 'Vancouver', 'Calgary', 'Edmonton','Ottawa', 'Winnipeg', 'Quebec', 'Hamilton']
data = []

# retrieve data from API
for city in cities:
  url = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={API_key}'
  response = requests.get(url)
  data.append(response.json())

# data preparation for the excel sheet
weather_data = []
for city_data in data:
  weather = {
    "City" :  city_data['name'],
    "temp" :  city_data['main']['temp'],
    "temp_min" : city_data['main']['temp_min'],
    "temp_max" : city_data['main']['temp_max'],
    "pressure" :  city_data['main']['pressure'],
    "humidity" :  city_data['main']['humidity'],
    "description" : city_data['weather'][0]['description'],
    "wind speed" :  city_data['wind']['speed'],
    "wind deg" : city_data['wind']['deg'],
    "clouds" : city_data['clouds']['all'],
    "Feels like" :  city_data['main']['feels_like'],
    "Visibility" : city_data['visibility']
  }
  weather_data.append(weather)


#######################################################################
# definition of weather related terms
# Visibility in weather refers to the distance one can see ahead in the atmosphere.
# wind degree describes the direction from which the wind is blowing.
#########################################################################

# create a virtual excel workbook
wb = openpyxl.Workbook()
worksheet = wb.active
worksheet.title = "Weather Data"

# create pandas dataframe
df = pd.DataFrame(weather_data)
#print(df)

# write column headers
worksheet.append([
  "City", "temp", "temp_min", "temp_max", "pressure", "humidity",
  "description", "wind speed", "wind deg", "clouds", "Feels like",
  "Visibility"])

#writing data to excel
for r, row in df.iterrows():
  worksheet.append([
    row['City'],  # City
    row['temp'],  # Temperature
    row['temp_min'],  # Minimum Temperature
    row['temp_max'],  # Maximum Temperature
    row['pressure'],  # Pressure
    row['humidity'],  # Humidity
    row['description'],  # Description
    row['wind speed'],  # Wind Speed
    row['wind deg'],  # Wind Degree
    row['clouds'],  # Clouds
    row['Feels like'],  # Feels Like
    row['Visibility']  # Visibility  
])

############# creation of excel charts ###########################
############## 01. Bar Chart #####################################
# data range definition
refObj1 = openpyxl.chart.Reference(worksheet=worksheet, min_col=1, min_row=2, max_col=2,max_row=10)
# This range will be used as the data source for the chart.

# create the bar chart
chartObj1 = openpyxl.chart.BarChart()
chartObj1.title = 'Temperature of Cities'

# add data from reference to the chart
chartObj1.add_data(refObj1, titles_from_data=False) # Add data without using titles from data

# Set x-axis labels
labels1 = openpyxl.chart.Reference(worksheet, min_col=1, min_row=2, max_row=10, max_col=1)
chartObj1.set_categories(labels1)

#set axis titles
chartObj1.x_axis.title = 'City'
chartObj1.y_axis.title = 'Temperature (K)'

chartObj1.legend = None  # Hide the legend or set it to 'False' to remove all legends

# drawing the chart starting from cell P2
worksheet.add_chart(chartObj1, 'P2')

################## 02. Line Chart ###############################################
refObj2 = openpyxl.chart.Reference(worksheet=worksheet, min_col=1, min_row=2, max_col=1,max_row=10)
refObj3 = openpyxl.chart.Reference(worksheet=worksheet, min_col=3, min_row=2, max_col=3,max_row=10)
refObj4 = openpyxl.chart.Reference(worksheet=worksheet, min_col=4, min_row=2, max_col=4,max_row=10)
# This range will be used as the data source for the chart.

# create the line chart
chartObj2 = openpyxl.chart.LineChart()
chartObj2.title = 'Min and Max Temperature variation of Cities'

# define series for the chart
series1 = openpyxl.chart.Series(refObj2, title='City')
series2 = openpyxl.chart.Series(refObj3, title='Min Temp')
series3 = openpyxl.chart.Series(refObj4, title='Max Temp')

# add series to the chart
#chartObj2.append(series1)
chartObj2.append(series2)
chartObj2.append(series3)

# Set x-axis labels
labels2 = openpyxl.chart.Reference(worksheet, min_col=1, min_row=2, max_row=10, max_col=1)
chartObj2.set_categories(labels2)

# set axis titles
chartObj2.x_axis.title = 'City'
chartObj2.y_axis.title = 'Temperature'

# drawing the chart starting from cell AA2
worksheet.add_chart(chartObj2, 'AA2')

########################## 03. Scatter Chart #################################

# defining X ,Y values for the chart
xvalues = openpyxl.chart.Reference(worksheet=worksheet, min_col=1, min_row=2, max_col=1,max_row=10)
yvalues= openpyxl.chart.Reference(worksheet=worksheet, min_col=8, min_row=2, max_col=8,max_row=10)
# This range will be used as the data source for the chart.
# create the scatter chart
chartObj3 = openpyxl.chart.ScatterChart()
chartObj3.title = 'Wind Speed distribution of cities'

# create a series for the chart
series4 = openpyxl.chart.Series( yvalues, xvalues)

# add the series to the chart
chartObj3.series.append(series4)

chartObj3.legend = None  # Hide the legend 

# set axis titles
chartObj3.x_axis.title = 'City'
chartObj3.y_axis.title = 'Wind Speed'
chartObj3.style = 10 # set chart style

# drawing the chart starting from cell P21
worksheet.add_chart(chartObj3, 'P21')

############# 04. Pie Chart ###########################################

# defining references for the chart
refObj2 = openpyxl.chart.Reference(worksheet=worksheet, min_col=1, min_row=2, max_col=1,max_row=10)
refObj6 = openpyxl.chart.Reference(worksheet=worksheet, min_col=6, min_row=2, max_col=6,max_row=10)

# create the pie chart
chartObj4 = openpyxl.chart.PieChart()
chartObj4.title = 'Humidity of cities'

# add data to the chart
chartObj4.add_data(refObj6, titles_from_data=False)
# set catergories 
chartObj4.set_categories(refObj2)

# drawing the chart starting from cell AA21
worksheet.add_chart(chartObj4,'AA21')

#########################################################################################
# Perform basic data analysis

mean_temp = df['temp'].mean()
median_humidity = df['humidity'].median()
std_dev_speed = df['wind speed'].std()

# print results for debugging purposes
print(mean_temp)
print(median_humidity)
print(std_dev_speed)

# Write analysis results to the sheet
worksheet.cell(row=17, column=1, value="Mean Temperature (C)")
worksheet.cell(row=17, column=2, value= mean_temp)
worksheet.cell(row=18, column=1, value="Median Humidity (%)")
worksheet.cell(row=18, column=2, value=median_humidity)
worksheet.cell(row=19, column=1, value="Standard Deviation of wind speed")
worksheet.cell(row=19, column=2, value= std_dev_speed)


# Save the workbook
wb.save("excel/weather_data.xlsx")

